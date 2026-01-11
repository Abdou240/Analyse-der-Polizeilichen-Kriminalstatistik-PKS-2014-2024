"""
Improved data extraction script for PKS Länder‑Fälle tables (2012–2024).

This module provides a flexible parser capable of handling the
structural variations present in the Excel tables for the years
2012–2024.  It detects header rows using multiple synonyms for
"Bundesland" (and its abbreviations), identifies the key, case and
AQ columns using configurable lists of patterns, and uses numeric
heuristics when column names are ambiguous.  Rows referring to the
whole of Germany ("Bundesrepublik Deutschland") or "Bund echte
Zählung der Tatverdächtigen" are excluded.  Note: the 2012 file in
the PKS archive does not contain a Länderaufschlüsselung (only
national totals), so 2012 will not appear in the final output.

The script writes two consolidated CSV files covering the years
2012–2024 (inclusive), one with absolute case counts and the other
with Aufklärungsquoten (AQ) rounded to one decimal place.  If
individual years are missing from the source data, they simply do not
appear in the final CSVs.

Usage:

    python improved_main.py

This will generate ``data/laender_relevant_cases_2012_2024.csv`` and
``data/laender_relevant_aq_2012_2024.csv`` in the ``data`` directory
relative to this script.
"""

import os
import re
from typing import List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook

# Define the offence keys of interest
KEY_MAP = {
    "diebstahl_insgesamt": "****00",
    "kfz_diebstahl": "***100",
    "wohnungseinbruch": "435*00",
}

# Synonyms for column headers.  All comparisons use case‑insensitive
# matching against whitespace‑normalised header strings.
KEY_SYNONYMS = [
    "schlüssel",
    "schl.",
    "schl",  # short form
    "strft",  # "Strft. Schl." as in 2013/2014
    "tatenschlüssel",
]

STATE_SYNONYMS = [
    # Standard German term
    "bundesland",
    # Abbreviations and alternative labels observed in various years
    "land",
    "bundesland/land",
    "bundesland land",
    "bundesl.",  # abbreviation used in some exports
    "bundesl",  # without punctuation
]

CASE_SYNONYMS = [
    # Generic column labels for case counts.  These strings may
    # appear alone or combined with a year (e.g. "erfasste Fälle 2013").
    "erfasste fälle",
    "anzahl erfasste fälle",
    "fallzahlen",
    "fälle",  # beware of "aufgeklärte Fälle"
    "anzahl",
    "anzahl fälle",
    # Include variants with explicit years to increase match likelihood.
    # "erfasste fälle 2012",
    # "erfasste fälle 2013",
    "erfasste fälle 2014",
    "erfasste fälle 2015",
    "erfasste fälle 2016",
    "erfasste fälle 2017",
    "erfasste fälle 2018",
    "erfasste fälle 2019",
    "erfasste fälle 2020",
    "erfasste fälle 2021",
    "erfasste fälle 2022",
    "erfasste fälle 2023",
    "erfasste fälle 2024",
]

AQ_SYNONYMS = [
    "aq",
    "aq in %",
    "aq%",
    "aufklärungsquote",
    "in % (aq)",
]


def normalize_text(value: object) -> str:
    """Convert a cell value to a normalised lowercase string.

    Newlines and multiple spaces are collapsed, leading/trailing
    whitespace is stripped, and the result is converted to lower case.
    Non‑string values are converted via str().  None becomes an empty
    string.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def to_float_de(value: object) -> float:
    """Parse a German/European formatted number to float.

    Handles thousands separators and decimal commas.  Dashes denote
    missing values.  Returns NaN for non‑numeric inputs.
    """
    if value is None:
        return float("nan")
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    s = str(value).strip()
    if s in {"-", "—", ""}:
        return float("nan")
    # remove spaces
    s = s.replace(" ", "")
    # decimal comma
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # remove thousand separators
        if re.search(r"\d\.\d{3}", s):
            s = s.replace(".", "")
    try:
        return float(s)
    except Exception:
        return float("nan")


def find_header_row(rows: List[List[object]]) -> Optional[int]:
    """Return the first row index containing a state identifier.

    Searches for any of the STATE_SYNONYMS in the row values.  Returns
    None if no such row is found.
    """
    for idx, row in enumerate(rows):
        norm = [normalize_text(cell) for cell in row]
        if any(any(syn in cell for syn in STATE_SYNONYMS) for cell in norm):
            return idx
    return None


def find_column(header: List[str], synonyms: List[str]) -> Optional[int]:
    """Return the column index matching any of the provided synonyms.

    Each entry in ``header`` should already be lower‑cased and
    whitespace‑normalised.
    """
    for idx, cell in enumerate(header):
        for syn in synonyms:
            if syn in cell:
                return idx
    return None


def find_aq_column(rows: List[List[object]], header_idx: int, max_scan: int = 6) -> Optional[int]:
    """Search for an AQ column in rows just below the header.

    Looks across up to ``max_scan`` rows starting from ``header_idx`` and
    returns the first column index where a cell matches any of the
    ``AQ_SYNONYMS``.  The search is case‑insensitive.
    """
    for r_idx in range(header_idx, min(header_idx + max_scan, len(rows))):
        norm = [normalize_text(c) for c in rows[r_idx]]
        for c_idx, cell in enumerate(norm):
            if any(syn == cell or syn in cell for syn in AQ_SYNONYMS):
                return c_idx
    return None


def parse_excel_file(path: str) -> pd.DataFrame:
    """Parse a PKS Excel file and extract relevant rows.

    Returns a DataFrame with columns: year, bundesland, key, cases,
    aq_percent.  Raises a ValueError if the file cannot be parsed.
    """
    # Determine year from filename
    m = re.search(r"_(\d{4})\.xlsx$", path)
    if not m:
        raise ValueError(f"Cannot determine year from {path}")
    year = int(m.group(1))
    wb = load_workbook(path, read_only=True, data_only=True)
    # iterate sheets looking for a header row
    target_rows = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = find_header_row(rows)
        if hdr_idx is not None:
            target_rows = rows
            header_idx = hdr_idx
            break
    if target_rows is None:
        raise ValueError(f"Header row not found in {path}")
    # Construct normalised header (combine header row and the next row)
    header_row = [normalize_text(c) for c in target_rows[header_idx]]
    # Some tables split header across two rows (e.g. 'AQ' appears one row below).
    # Merge with the second header row by appending non‑blank cells.
    merged_header = header_row.copy()
    if header_idx + 1 < len(target_rows):
        next_row = [normalize_text(c) for c in target_rows[header_idx + 1]]
        for idx in range(min(len(merged_header), len(next_row))):
            if next_row[idx] and next_row[idx] not in merged_header[idx]:
                # combine with a space if both are non‑empty
                if merged_header[idx]:
                    merged_header[idx] = merged_header[idx] + " " + next_row[idx]
                else:
                    merged_header[idx] = next_row[idx]
    # locate key, state, cases and AQ columns
    idx_state = find_column(merged_header, STATE_SYNONYMS)
    if idx_state is None:
        raise ValueError(f"State column not found in {path}")
    idx_key = find_column(merged_header, KEY_SYNONYMS)
    if idx_key is None:
        raise ValueError(f"Key column not found in {path}")
    # Cases: prefer columns with explicit cases synonyms but avoid 'versuche', 'aufgeklärte', etc.
    idx_cases = find_column(merged_header, CASE_SYNONYMS)
    # Validate that the found column is not a 'versuche' column
    if idx_cases is not None:
        col_name = merged_header[idx_cases]
        if 'versuch' in col_name or 'versuche' in col_name or 'aufgekl' in col_name:
            idx_cases = None
    # If still None, fall back to numeric detection as in the previous script
    if idx_cases is None:
        for col in range(idx_state + 1, len(merged_header)):
            # sample numeric content
            vals = [target_rows[header_idx + 2 + r][col] for r in range(min(25, len(target_rows) - header_idx - 2))
                    if col < len(target_rows[header_idx + 2 + r]) and target_rows[header_idx + 2 + r][col] is not None]
            if not vals:
                continue
            numish = sum(
                (isinstance(v, (int, float)) and not pd.isna(v)) or bool(re.search(r"\d", str(v)))
                for v in vals
            )
            if numish >= 15:
                idx_cases = col
                break
        if idx_cases is None:
            raise ValueError(f"Cases column could not be inferred in {path}")
    # AQ column detection
    idx_aq = find_aq_column(target_rows, header_idx)
    if idx_aq is None:
        # attempt to match via header synonyms as fallback
        idx_aq = find_column(merged_header, AQ_SYNONYMS)
    if idx_aq is None:
        raise ValueError(f"AQ column not found in {path}")
    # Extract data
    records: List[Tuple[int, str, str, float, float]] = []
    for row in target_rows[header_idx + 2:]:
        # ensure row length
        if idx_key >= len(row) or row[idx_key] is None:
            continue
        raw_key = normalize_text(row[idx_key]).upper()
        # Only consider 6‑character keys consisting of digits or '*'
        if not re.fullmatch(r"[0-9\*]{6}", raw_key):
            continue
        # state
        state = normalize_text(row[idx_state]).title() if idx_state < len(row) else ""
        # exclude national totals
        if state.lower() in {"bundesrepublik deutschland", "bund echte zählung der tatverdächtigen"}:
            continue
        cases = to_float_de(row[idx_cases]) if idx_cases < len(row) else float("nan")
        aq = to_float_de(row[idx_aq]) if idx_aq < len(row) else float("nan")
        records.append((year, state, raw_key, cases, aq))
    df = pd.DataFrame(records, columns=["year", "bundesland", "key", "cases", "aq_percent"])
    return df


def process_all_files(files: List[str]) -> pd.DataFrame:
    """Parse all files and return a concatenated DataFrame."""
    data_frames: List[pd.DataFrame] = []
    for path in files:
        try:
            df = parse_excel_file(path)
        except Exception as e:
            # Log the failure and continue
            print(f"Warning: {os.path.basename(path)}: {e}")
            continue
        data_frames.append(df)
    if not data_frames:
        raise RuntimeError("No data frames parsed")
    return pd.concat(data_frames, ignore_index=True)


def main() -> None:
    # Directory containing the converted Excel files (must exist)
    base_dir = os.path.dirname(__file__)
    data_dir = os.path.join(base_dir, "data_converted")
    if not os.path.isdir(data_dir):
        raise RuntimeError(f"Data directory not found: {data_dir}")
    # List all years' files
    files = sorted(
        os.path.join(data_dir, f)
        for f in os.listdir(data_dir)
        if f.lower().endswith(".xlsx") and re.search(r"_(201[4-9]|202[0-4])\.xlsx$", f)
       # if f.lower().endswith(".xlsx") and re.search(r"_\d{4}\.xlsx$", f)
    )
    # Parse files
    panel = process_all_files(files)
    # Keep only relevant keys
    panel = panel[panel["key"].isin(KEY_MAP.values())].copy()
    # Pivot to wide format for cases
    cases_wide = (
        panel.pivot_table(index=["year", "bundesland"], columns="key", values="cases", aggfunc="first")
        .reset_index()
        .rename(columns={v: k for k, v in KEY_MAP.items()})
    )
    cases_wide = cases_wide[["year", "bundesland", "diebstahl_insgesamt", "kfz_diebstahl", "wohnungseinbruch"]]
    cases_wide = cases_wide.sort_values(["year", "bundesland"])
    # Pivot for AQ and round
    aq_wide = (
        panel.pivot_table(index=["year", "bundesland"], columns="key", values="aq_percent", aggfunc="first")
        .reset_index()
        .rename(columns={v: k for k, v in KEY_MAP.items()})
    )
    aq_wide = aq_wide[["year", "bundesland", "diebstahl_insgesamt", "kfz_diebstahl", "wohnungseinbruch"]]
    aq_wide[["diebstahl_insgesamt", "kfz_diebstahl", "wohnungseinbruch"]] = aq_wide[
        ["diebstahl_insgesamt", "kfz_diebstahl", "wohnungseinbruch"]
    ].round(1)
    aq_wide = aq_wide.sort_values(["year", "bundesland"])
    # Write outputs
    out_dir = os.path.join(base_dir, "data")
    os.makedirs(out_dir, exist_ok=True)
    cases_csv = os.path.join(out_dir, "laender_relevant_cases_2014_2024.csv")
    aq_csv = os.path.join(out_dir, "laender_relevant_aq_2014_2024.csv")
    cases_wide.to_csv(cases_csv, index=False)
    aq_wide.to_csv(aq_csv, index=False)
    print(f"Wrote {cases_csv}")
    print(f"Wrote {aq_csv}")
    # Validate presence of years
    present_years = sorted(cases_wide['year'].unique().tolist())
    print(f"Years present in output: {present_years}")


if __name__ == "__main__":
    main()